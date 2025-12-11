import cv2
import pytesseract
import re
import pandas as pd
import os
import glob
from datetime import datetime, timedelta
from openpyxl import load_workbook
from typing import List, Dict, Any, Union

# Configuración
ARCHIVO_EXCEL = 'Datos_PedidosYa.xlsx'
HOJA_NOMBRE = 'Pedidos'
CARPETA_FOTOS = 'capturas'

# Windows config
pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def normalizar_para_comparar(valor: Union[str, datetime, float]) -> str:
    """
    Estandariza fechas y textos a un formato de string único para evitar duplicados falsos.
    Convierte objetos datetime o strings variados al formato 'DD/MM/AAAA HH:MM'.

    Args:
        valor (Union[str, datetime, float]): El valor de la fecha u hora proveniente de Excel (datetime/str) u OCR.

    Returns:
        str: La fecha formateada como string "dd/mm/yyyy HH:MM" o cadena vacía si el input no es válido.
    """
    if pd.isna(valor) or valor == "": return ""
    if isinstance(valor, datetime): return valor.strftime("%d/%m/%Y %H:%M")
    
    texto = str(valor).strip()
    formatos = ["%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%Y-%m-%d %H:%M"]
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).strftime("%d/%m/%Y %H:%M")
        except ValueError: continue
    return texto

def es_separador(texto: str) -> bool:
    """
    Identifica si una línea de texto es 'basura' técnica, metadata o encabezados de la app,
    y no parte del nombre real de un local.

    Args:
        texto (str): La línea de texto extraída por el OCR a analizar.

    Returns:
        bool: True si la línea es un separador/metadata, False si es contenido útil (nombre potencial).
    """
    t = texto.lower().strip()
    # IDs numéricos largos (ej: 1811824615)
    if re.match(r'^\d{7,}$', t): return True
    # Estados y textos de app
    if "pedido agrupado" in t: return True
    if "completado" in t: return True
    if "cancelado" in t: return True
    if "ver detalles" in t: return True
    # Encabezados de fecha (ej: "semana 49", "vie, 5")
    if "semana" in t and re.search(r'\d{2}', t): return True
    if re.match(r'^\w{3}, \d+', t): return True
    # Información de pagos/dinero
    if "promedio" in t or "ars" in t: return True
    if "horas conectado" in t: return True
    return False

def obtener_diccionario_locales(df_historico: pd.DataFrame) -> Dict[str, Dict[str, str]]:
    """
    Crea un diccionario de referencia ('memoria') con los datos de locales ya visitados.
    Solo aprende de registros históricos que tengan tanto el Nombre como la Dirección completos.

    Args:
        df_historico (pd.DataFrame): El DataFrame con los datos históricos leídos del Excel.

    Returns:
        Dict[str, Dict[str, str]]: Un diccionario donde la clave es el nombre del local y el valor 
        es otro diccionario con 'Direccion_Local', 'Tipo_Negocio' y 'Cadena'.
    """
    conocimiento = {}
    if df_historico.empty: return conocimiento
    cols = ['Nombre_Local', 'Direccion_Local', 'Tipo_Negocio', 'Cadena']
    for c in cols: 
        if c not in df_historico.columns: return {}

    # Filtramos para quedarnos solo con filas que tengan info completa (estrategia conservadora)
    df_util = df_historico.dropna(subset=['Nombre_Local', 'Direccion_Local'])
    
    for local in df_util['Nombre_Local'].unique():
        # Tomamos la última aparición de este local
        datos = df_util[df_util['Nombre_Local'] == local].iloc[-1]
        conocimiento[local] = {
            'Direccion_Local': datos['Direccion_Local'],
            'Tipo_Negocio': datos.get('Tipo_Negocio', ''),
            'Cadena': datos.get('Cadena', '')
        }
    return conocimiento

def procesar_imagen_ocr(ruta_imagen: str) -> List[Dict[str, Any]]:
    """
    Procesa una imagen individual usando OCR para extraer una lista de pedidos.
    Aplica lógica de limpieza, detección de fechas/horas y reconstrucción de nombres multilínea.

    Args:
        ruta_imagen (str): La ruta del archivo de imagen a procesar.

    Returns:
        List[Dict[str, Any]]: Una lista de diccionarios, donde cada diccionario representa un pedido 
        con claves 'Hora_Aceptacion', 'Hora_Entrega' y 'Nombre_Local'.
    """
    img = cv2.imread(ruta_imagen)
    if img is None: return []
    
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # Ajuste: a veces menos blur ayuda a leer números entre paréntesis
    # gray = cv2.GaussianBlur(gray, (3, 3), 0) 
    _, thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    texto = pytesseract.image_to_string(thresh, lang='spa')
    # Filtramos líneas vacías
    lineas = [l.strip() for l in texto.split('\n') if l.strip()]

    meses = {'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04', 'may': '05', 'jun': '06',
             'jul': '07', 'ago': '08', 'sep': '09', 'oct': '10', 'nov': '11', 'dic': '12'}
    
    fecha_base_str, anio_base = None, None
    pedidos_en_foto = []

    # Regex
    regex_fecha = r"(\w{3}),\s*(\d{1,2})\s*de\s*(\w{3})"
    regex_anio = r"20\d{2}"
    regex_horas = r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})"

    # 1. Contexto Temporal
    for linea in lineas:
        if not fecha_base_str:
            m = re.search(regex_fecha, linea, re.IGNORECASE)
            if m: fecha_base_str = f"{m.group(2).zfill(2)}/{meses.get(m.group(3).lower()[:3], '01')}"
        if not anio_base:
            m = re.search(regex_anio, linea)
            if m: anio_base = m.group(0)
    
    if not anio_base: anio_base = str(datetime.now().year)

    # 2. Extracción de Pedidos
    if fecha_base_str:
        formato_fecha_obj = "%d/%m/%Y"
        try:
            fecha_obj_inicial = datetime.strptime(f"{fecha_base_str}/{anio_base}", formato_fecha_obj)
        except ValueError: return []

        for i, linea in enumerate(lineas):
            match_horas = re.search(regex_horas, linea)
            if match_horas:
                # Extraer horas
                h_inicio_str, h_fin_str = match_horas.group(1), match_horas.group(2)
                try:
                    h_ini = datetime.strptime(h_inicio_str, "%H:%M").time()
                    h_fin = datetime.strptime(h_fin_str, "%H:%M").time()
                except ValueError: continue

                dt_acep = datetime.combine(fecha_obj_inicial.date(), h_ini)
                dt_entr = datetime.combine(fecha_obj_inicial.date(), h_fin)

                if dt_entr < dt_acep: dt_entr += timedelta(days=1)

                # --- LÓGICA V6: DETECCIÓN DE NOMBRE INTELIGENTE ---
                if i > 0:
                    nombre = lineas[i-1]
                    
                    # Miramos una línea más arriba (i-2)
                    if i > 1:
                        linea_anterior = lineas[i-2]
                        # Si la línea anterior NO es un separador (ID, fecha, etc), 
                        # asumimos que es la primera parte del nombre.
                        if not es_separador(linea_anterior):
                            nombre = f"{linea_anterior} {nombre}"
                else:
                    nombre = "Desconocido"

                # Limpieza final del nombre
                nombre = re.sub(r'^[\W_]+', '', nombre) # Sacar símbolos al inicio
                nombre = re.sub(r'[^\w\s\(\)]+$', '', nombre) # Sacar basura al final (menos paréntesis)
                nombre = nombre.strip()

                pedidos_en_foto.append({
                    'Hora_Aceptacion': dt_acep.strftime("%d/%m/%Y %H:%M"),
                    'Hora_Entrega': dt_entr.strftime("%d/%m/%Y %H:%M"),
                    'Nombre_Local': nombre
                })
    return pedidos_en_foto

def rellenar_huecos_historicos(archivo: str, hoja: str, diccionario_locales: Dict[str, Dict[str, str]]) -> None:
    """
    Recorre el archivo Excel y rellena las celdas vacías (Dirección, Tipo, Cadena) 
    si encuentra coincidencias en el diccionario de locales conocidos.

    Args:
        archivo (str): Ruta al archivo Excel.
        hoja (str): Nombre de la hoja a editar.
        diccionario_locales (Dict[str, Dict[str, str]]): Diccionario con la información aprendida de los locales.

    Returns:
        None: Modifica el archivo Excel in-place y lo guarda.
    """
    print(">>> 🧹 Iniciando limpieza y autocompletado...")
    book = load_workbook(archivo)
    sheet = book[hoja]
    cambios = 0
    
    COL_NOMBRE, COL_DIR, COL_TIPO, COL_CADENA = 5, 6, 7, 8
    
    for row in range(2, sheet.max_row + 1):
        nombre_celda = sheet.cell(row=row, column=COL_NOMBRE).value
        if nombre_celda:
            nombre_key = str(nombre_celda).strip()
            if nombre_key in diccionario_locales:
                datos = diccionario_locales[nombre_key]
                
                # Función auxiliar para rellenar
                def rellenar(col, valor):
                    c = sheet.cell(row=row, column=col)
                    if not c.value and valor: c.value = valor; return 1
                    return 0
                
                cambios += rellenar(COL_DIR, datos['Direccion_Local'])
                cambios += rellenar(COL_TIPO, datos['Tipo_Negocio'])
                cambios += rellenar(COL_CADENA, datos['Cadena'])

    if cambios > 0:
        book.save(archivo)
        print(f"✅ Se rellenaron {cambios} celdas.")
    else:
        print("✓ No hubo cambios necesarios.")

def main() -> None:
    """
    Función principal que orquesta el flujo ETL:
    1. Carga histórico y aprende locales.
    2. Procesa imágenes nuevas.
    3. Filtra duplicados y nuevos registros.
    4. Inserta en Excel.
    5. Ejecuta limpieza y autocompletado.

    Returns:
        None
    """
    if not os.path.exists(ARCHIVO_EXCEL):
        print(f"❌ Falta archivo: {ARCHIVO_EXCEL}"); return

    print("--- LÓGICA CONTEXTUAL ---")
    try:
        df_hist = pd.read_excel(ARCHIVO_EXCEL, sheet_name=HOJA_NOMBRE)
    except Exception as e: print(f"Error Excel: {e}"); return

    dicc = obtener_diccionario_locales(df_hist)
    
    # IDs y Firmas
    if 'ID_Pedido' in df_hist.columns and not df_hist.empty:
        df_hist['ID_Pedido'] = pd.to_numeric(df_hist['ID_Pedido'], errors='coerce')
        ultimo_id = int(df_hist['ID_Pedido'].max()) if not df_hist['ID_Pedido'].isnull().all() else 0
    else: ultimo_id = 0

    firmas = set()
    if not df_hist.empty:
        for _, r in df_hist.iterrows():
            ta, te = normalizar_para_comparar(r.get('Hora_Aceptacion')), normalizar_para_comparar(r.get('Hora_Entrega'))
            if ta and te: firmas.add(f"{ta}_{te}")

    # Procesar
    imgs = [f for f in glob.glob(os.path.join(CARPETA_FOTOS, '*.*')) if f.lower().endswith(('.png','.jpg','.jpeg'))]
    print(f"📸 Fotos encontradas: {len(imgs)}")

    nuevos, contador = [], ultimo_id + 1

    for ruta in imgs:
        for p in procesar_imagen_ocr(ruta):
            firma = f"{p['Hora_Aceptacion']}_{p['Hora_Entrega']}"
            if firma in firmas: continue

            nuevos.append({
                'ID_Pedido': contador, 'ID_Turno': '', 
                'Hora_Aceptacion': p['Hora_Aceptacion'], 'Hora_Entrega': p['Hora_Entrega'], 
                'Nombre_Local': p['Nombre_Local'],
                'Direccion_Local': '', 'Tipo_Negocio': '', 'Cadena': '', 'Direccion_Cliente': '', 'Propina_Pedido': ''
            })
            firmas.add(firma); contador += 1

    # Insertar
    if nuevos:
        print(f"📝 Insertando {len(nuevos)} pedidos...")
        book = load_workbook(ARCHIVO_EXCEL)
        sheet = book[HOJA_NOMBRE]
        start = sheet.max_row + 1
        cols = ['ID_Pedido', 'ID_Turno', 'Hora_Aceptacion', 'Hora_Entrega', 'Nombre_Local', 'Direccion_Local', 'Tipo_Negocio', 'Cadena', 'Direccion_Cliente', 'Propina_Pedido']
        
        for i, reg in enumerate(nuevos):
            for j, k in enumerate(cols, 1):
                sheet.cell(row=start+i, column=j, value=reg[k])
        
        book.save(ARCHIVO_EXCEL)
        print("✅ Guardado.")
    else: print("⚠ No hay pedidos nuevos.")

    rellenar_huecos_historicos(ARCHIVO_EXCEL, HOJA_NOMBRE, dicc)
    print("--- FIN ---")

if __name__ == "__main__":
    main()