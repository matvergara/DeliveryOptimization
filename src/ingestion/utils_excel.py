import pandas as pd
from typing import Dict, Any
from openpyxl import load_workbook


def leer_hoja_raw(ruta_excel: str, hoja: str) -> pd.DataFrame:
    """
    Lee una hoja del archivo Excel RAW.

    Args:
        ruta_excel: Ruta al archivo Excel.
        hoja: Nombre de la hoja a leer.

    Returns:
        Datos crudos de la hoja.
    """
    return pd.read_excel(ruta_excel, sheet_name=hoja)


def obtener_siguiente_id(df: pd.DataFrame, columna_id: str) -> int:
    """
    Obtiene el siguiente ID incremental para una tabla RAW.

    Args:
        df: DataFrame histórico.
        columna_id: Nombre de la columna ID.

    Returns:
        ID siguiente.
    """
    if df.empty or columna_id not in df.columns:
        return 1

    return int(df[columna_id].max()) + 1


def insertar_registro_raw(ruta_excel: str, hoja: str, registro: Dict[str, Any]) -> None:
    """
    Inserta un nuevo registro al final de una hoja RAW.

    Args:
        ruta_excel: Ruta al archivo Excel.
        hoja: Nombre de la hoja.
        registro: Registro a insertar.
            Las claves deben respetar el orden de columnas del Excel.
    """
    book = load_workbook(ruta_excel)
    sheet = book[hoja]

    fila = sheet.max_row + 1

    for col_idx, valor in enumerate(registro.values(), start=1):
        sheet.cell(row=fila, column=col_idx, value=valor)

    book.save(ruta_excel)
